import datetime
from os import path
from tkinter import ttk
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
import customtkinter as s
data = datetime.date.today()
data= data.strftime("%d/%m/%Y")

# Variáveis globais
diretorio = path.join(path.expanduser("~"), "Documents\\data.xlsx")
janelaPesquisa = None
tree = None

def pesquisa(nomePesquisado):
    try:
        # Carregar dados do Excel
        df = pd.read_excel(diretorio, sheet_name='Sheet')
        df2 = df[df['nome'].str.contains(nomePesquisado)]

        # Esconder janela principal e criar nova janela de pesquisa
        janelaMenu.withdraw()
        JanelaPesquisa = s.CTkToplevel(janelaMenu)
        JanelaPesquisa.geometry("1000x600")
        JanelaPesquisa.title("Clientes Cadastrados")
        JanelaPesquisa.resizable(False, False)
        s.set_appearance_mode("Dark")

        # Função para converter dataframe em Treeview
        def dataframe_to_tk_treeview(root, dataframe):
            container = ttk.Frame(root)
            container.pack(fill='both', expand=True)

            # Criar o Treeview
            tree = ttk.Treeview(container)
            tree.pack(fill='both', expand=True)

            # Configurar colunas
            tree["columns"] = list(dataframe.columns)
            tree["show"] = "headings"

            # Adicionar cabeçalhos
            for column in dataframe.columns:
                tree.heading(column, text=column)

            # Adicionar dados
            for index, row in dataframe.iterrows():
                tree.insert("", "end", values=list(row))

            # Botão de voltar
            BTNVoltar = s.CTkButton(JanelaPesquisa, text="Voltar", command=lambda: Voltar(JanelaPesquisa))
            BTNVoltar.pack(padx=5, pady=5)

            return tree

        # Chamar função para criar Treeview
        dataframe_to_tk_treeview(JanelaPesquisa, df2)

    except Exception as e:
        statusvar.set("Ainda não há cadastros.")

def Voltar(j):
    janelaMenu.deiconify()
    EntryNome.delete(0, "end")
    EntryNome.focus()
    j.withdraw()

def EnvioExcel(nome, servicos, dataVenda, pago, formaPagamento, telefone):
    statusvar.set("")  # Limpar status
    try:
        # Carregar ou criar arquivo Excel
        book = load_workbook(diretorio)
        sheet = book['Sheet']
        statusvar.set("Tabela encontrada")

    except:
        # Se não encontrar, criar novo arquivo Excel
        clientes = ["nome", "servicos", "dataVenda", "pago", "formaPagamento", "telefone"]
        book = openpyxl.Workbook()
        sheet = book['Sheet']
        sheet.append(clientes)
        book.save(diretorio)
        statusvar.set("Nova tabela criada")
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 11
        sheet.column_dimensions['F'].width = 12

    # Verificar duplicidade e salvar dados
    df = pd.read_excel(diretorio, sheet_name='Sheet')
    if nome in df["nome"].values and dataVenda in df['dataVenda']:
        statusvar.set("Cadastro repetido")
    elif nome == "" or pago == "":
        statusvar.set("Falta algo!")
    else:
        try:
            sheet.append([nome, servicos, dataVenda, pago, formaPagamento, telefone])
            book.save(diretorio)
            statusvar.set("Cadastrado com sucesso!")
        except:
            statusvar.set("Feche o Excel!")

    # Limpar campos após salvar
    EntryNome.delete(0, "end")
    EntryValor.delete(0, "end")
    EntryNumero.delete(0, "end")
    EntryOBS.delete('1.0', s.END)
    EntryNome.focus()

# Interface gráfica
janelaMenu = s.CTk()
janelaMenu.geometry("630x250")
janelaMenu.title("Cadastro de Serviços")
janelaMenu.resizable(False, False)
s.set_appearance_mode("Dark")
s.set_default_color_theme("dark-blue")

statusvar = s.StringVar()
statusvar.set("Conectado!")

#Variavel de controle 
LabelCtrl = s.CTkLabel(janelaMenu, 
                       textvariable=statusvar)
LabelCtrl.grid(row=6, column=2, padx=10)

#Label "Nome"
LabelNome = s.CTkLabel(janelaMenu, 
                       text="Nome:")
LabelNome.grid(row=1, column=1, pady=10, padx=14)

#Caixa de entrada "Nome"
EntryNome = s.CTkEntry(janelaMenu,
                       placeholder_text="Digite o nome",
                       border_color="Pink", 
                       width=250)
EntryNome.grid(row=1, column=2, pady=10, sticky="W")

#Label "Valor"
LabelValor = s.CTkLabel(janelaMenu,
                        text="Valor:")
LabelValor.grid(row=1, column=3, padx=10)

#Caixa de entrada "Valor"
EntryValor = s.CTkEntry(janelaMenu,
                        placeholder_text="Valor...",
                        border_color="Pink",
                        width=100)
EntryValor.grid(row=1, column=4, pady=10, sticky="W")

#Label "Celular"
LabelNumero = s.CTkLabel(janelaMenu,
                         text="Celular:")
LabelNumero.grid(row=2, column=1)

#Caixa de entrada "Numero"
EntryNumero = s.CTkEntry(janelaMenu,
                         placeholder_text="(99)99999-9999",
                         border_color="Pink")
EntryNumero.grid(row=2, column=2, sticky="W", padx=5, pady=10)

#Caixa de opção de forma de pagamento
OPTPagamento = s.CTkOptionMenu(janelaMenu,
                               text_color="Black", 
                               fg_color="Pink",
                               button_color="Pink",
                               button_hover_color="Pink",
                               values=["Outro", "Cartão", "Dinheiro", "Pix"],
                               width=85)
OPTPagamento.grid(row=1, column=5, padx=10)

#Label "Serviços"
LabelOBS = s.CTkLabel(janelaMenu,
                      text="Serviços:")
LabelOBS.grid(row=3, column=1)

#Label data atual
LabelData = s.CTkLabel(janelaMenu, 
                       text=data)
LabelData.grid(row=6, column=1, padx=10, sticky="S")

#Caixa de entrada "Serviços"
EntryOBS = s.CTkTextbox(janelaMenu,
                        height=100,
                        width=300,
                        border_width=2,
                        border_color="Pink")
EntryOBS.grid(row=3, column=2, columnspan=2, rowspan=3, pady=10)

#Botão "Salvar"
BTNSalvar = s.CTkButton(janelaMenu,
                        text="Salvar",
                        text_color="Black",
                        fg_color="Pink",
                        width=100,
                        command=lambda: EnvioExcel(EntryNome.get(),
                                                   EntryOBS.get("1.0", "end-1c"),
                                                   data, EntryValor.get(),
                                                   OPTPagamento.get(),
                                                   EntryNumero.get()))
BTNSalvar.grid(row=6, column=5)

#Botão "Pesquisar" 
BTNVer = s.CTkButton(janelaMenu, text="Pesquisar",
                     text_color="Black", 
                     fg_color="Pink", 
                     command=lambda: pesquisa(EntryNome.get()), width=100)
BTNVer.grid(row=6, column=4, padx=5)

janelaMenu.mainloop()